#!/usr/bin/env python3
"""Read what sheetedit wrote with openpyxl, and check it means what sheetedit intended.

Round-trips prove the project agrees with itself. LibreOffice judges what it supports and drops
slicers entirely. The ECMA schemas judge structure but not meaning. This asks the remaining
question: does a SEPARATE implementation, in another language, sharing no code with this one,
understand the file the way we meant it?

The workbook and the expectations both come from src/adapters/xlsx/openpyxl-corpus.test.ts, so
what is asserted is what that test MEANT to write, not whatever it happened to produce.

Usage: openpyxl-check.py [corpus-dir]   (default .cache/openpyxl-corpus)
"""
import json
import sys
import warnings
from pathlib import Path

warnings.simplefilter("ignore")  # openpyxl warns about parts it does not model; not our concern

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed (see `npm run check:openpyxl`)", file=sys.stderr)
    sys.exit(2)

corpus = Path(sys.argv[1] if len(sys.argv) > 1 else ".cache/openpyxl-corpus")
book = corpus / "authored.xlsx"
spec = corpus / "expected.json"
if not book.exists() or not spec.exists():
    print(f"no corpus in {corpus}: run `npm run check:openpyxl`, which writes it first", file=sys.stderr)
    sys.exit(2)

want = json.loads(spec.read_text())
failures: list[str] = []


def check(label: str, got, expected) -> None:
    if got != expected:
        failures.append(f"{label}: openpyxl sees {got!r}, sheetedit meant {expected!r}")


wb = openpyxl.load_workbook(book)
check("sheet name", wb.sheetnames[0], want["sheet"])
ws = wb[want["sheet"]]

# The table (ListObject): the thing that makes a range a named, structured thing.
tables = {name: str(t.ref) if hasattr(t, "ref") else str(t) for name, t in ws.tables.items()}
if want["table"]["name"] not in tables:
    failures.append(f"table: openpyxl sees {list(tables)}, sheetedit meant {want['table']['name']!r}")
else:
    check("table ref", tables[want["table"]["name"]], want["table"]["ref"])

# Conditional formatting, by the range it covers.
seen_cf = {}
for rng in ws.conditional_formatting:
    for rule in rng.rules:
        seen_cf.setdefault(str(rng.sqref), []).append(rule)
for expect_cf in want["condFormats"]:
    rules = seen_cf.get(expect_cf["range"])
    if not rules:
        failures.append(f"conditional format {expect_cf['range']}: openpyxl sees none (it has {list(seen_cf)})")
        continue
    rule = rules[0]
    check(f"cf {expect_cf['range']} type", rule.type, expect_cf["type"])
    if "operator" in expect_cf:
        check(f"cf {expect_cf['range']} operator", rule.operator, expect_cf["operator"])
    if "formula" in expect_cf:
        check(f"cf {expect_cf['range']} formula", list(rule.formula)[0] if rule.formula else None, expect_cf["formula"])
    if "text" in expect_cf:
        check(f"cf {expect_cf['range']} text", rule.text, expect_cf["text"])

# Data validation, including the messages a rule uses to explain itself.
w = want["validation"]
dv = next((d for d in ws.data_validations.dataValidation if str(d.sqref) == w["range"]), None)
if dv is None:
    failures.append(f"validation {w['range']}: openpyxl sees {[str(d.sqref) for d in ws.data_validations.dataValidation]}")
else:
    check("dv type", dv.type, w["type"])
    check("dv operator", dv.operator, w["operator"])
    check("dv formula1", dv.formula1, w["formula1"])
    check("dv formula2", dv.formula2, w["formula2"])
    check("dv promptTitle", dv.promptTitle, w["promptTitle"])
    check("dv prompt", dv.prompt, w["prompt"])
    check("dv errorTitle", dv.errorTitle, w["errorTitle"])
    check("dv error", dv.error, w["error"])

# A hyperlink, a comment, the frozen pane and the merge.
link = ws[want["hyperlink"]["cell"]].hyperlink
check("hyperlink target", link.target if link else None, want["hyperlink"]["target"])
comment = ws[want["comment"]["cell"]].comment
check("comment text", (comment.text if comment else None), want["comment"]["text"])
check("freeze panes", ws.freeze_panes, want["freezePanes"])
check("merged range", want["merged"] in [str(r) for r in ws.merged_cells.ranges], True)

# And the values themselves, since a reader that sees the structure but not the data proves little.
for ref, value in want["values"].items():
    check(f"value {ref}", ws[ref].value, value)

if failures:
    print(f"openpyxl disagrees with sheetedit in {len(failures)} place(s):\n")
    for f in failures:
        print(f"  {f}")
    sys.exit(1)
print(f"openpyxl reads the authored workbook as sheetedit meant it ({len(want['values']) + 12} checks).")
